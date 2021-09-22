from io import BytesIO
from zipfile import BadZipFile
import datetime
from openpyxl import load_workbook

from . import utils


def handle_uploaded_file(upform, my_file, myClass):
    # pylint: disable=R0912
    try:
        my_wb = load_workbook(filename=BytesIO(my_file.read()), data_only=True)
    except BadZipFile:
        upform.add_error(
            "file",
            "Le fichier importé ne semble pas être du bon format, 'xlsx' est le format attendu",
        )
        return {"success": utils.ReturnStatus.ERROR}
    try:
        my_ws = my_wb[myClass.sheet_name]
    except KeyError:
        upform.add_error(
            "file",
            f"Le fichier importé doit avoir une feuille nommée '{myClass.sheet_name}'",
        )
        return {"success": utils.ReturnStatus.ERROR}
    import_warnings = []
    column_from_index = {}
    for col in my_ws.iter_cols(
        min_col=1, max_col=my_ws.max_column, min_row=1, max_row=1
    ):
        for cell in col:
            if cell.value is None:
                continue
            if cell.value not in myClass.import_mapping:
                import_warnings.append(
                    Exception(
                        f"La colonne nommée '{cell.value}' est inconnue, "
                        + "elle sera ignorée. Les colonnes attendues sont : "
                        + f"{', '.join(myClass.import_mapping.keys())}"
                    )
                )
                continue
            column_from_index[cell.column] = str(cell.value).strip()

    error_column = False
    for key in myClass.import_mapping:
        if key not in list(column_from_index.values()):
            upform.add_error(
                "file", f"Le fichier importé doit avoir une colonne nommée '{key}'"
            )
            error_column = True
    if error_column:
        return {"success": utils.ReturnStatus.ERROR}

    # transform each line into object
    my_objects, import_warnings = get_object_from_worksheet(
        my_ws, column_from_index, myClass, import_warnings
    )

    return {
        "success": utils.ReturnStatus.SUCCESS
        if len(import_warnings) == 0
        else utils.ReturnStatus.WARNING,
        "objects": my_objects,
        "import_warnings": import_warnings,
    }


def get_object_from_worksheet(my_ws, column_from_index, myClass, import_warnings):
    my_objects = []
    for row in my_ws.iter_rows(
        min_row=3, max_row=my_ws.max_row, min_col=1, max_col=my_ws.max_column
    ):
        my_row, empty_line, new_warnings = extract_row(
            row, column_from_index, myClass.import_mapping
        )
        import_warnings = [*import_warnings, *new_warnings]

        # Ignore if the line is empty
        if not empty_line:
            my_objects.append(my_row)
    return my_objects, import_warnings


def extract_row(row, column_from_index, import_mapping):
    # pylint: disable=R0912
    new_warnings = []
    my_row = {}
    empty_line = True
    for cell in row:
        # Ignore unknown column
        if cell.column not in column_from_index:
            continue

        # Check the empty lines to don't fill it
        if cell.value is not None:
            empty_line = False
        else:
            continue

        value = None
        model_field = import_mapping[column_from_index[cell.column]]

        if isinstance(model_field, str):
            key = model_field
            value = cell.value
        else:
            key = model_field.name

            # Date case
            if model_field.get_internal_type() == "DateField":
                if isinstance(cell.value, datetime.datetime):
                    value = utils.format_date_for_form(cell.value)
                else:
                    new_warnings.append(
                        Exception(
                            f"{cell.column_letter}{cell.row} : La valeur '{cell.value}' "
                            + f"de la colonne {column_from_index[cell.column]} "
                            + "doit être une date"
                        )
                    )

            # TextChoices case
            elif (
                model_field.get_internal_type() == "CharField"
                and model_field.choices is not None
            ):
                if cell.value is not None:
                    value = next(
                        (x[0] for x in model_field.choices if x[1] == cell.value), None
                    )
                    if (
                        value is None
                    ):  # value is not Null but not in the choices neither
                        new_warnings.append(
                            Exception(
                                f"{cell.column_letter}{cell.row} : La valeur '{cell.value}' "
                                + f"de la colonne {column_from_index[cell.column]} "
                                + "doit faire partie des valeurs : "
                                + f"{', '.join(map(lambda x : x[1], model_field.choices))}"
                            )
                        )

            # Float case
            elif model_field.get_internal_type() == "FloatField":
                if cell.value is not None:
                    if isinstance(cell.value, (float, int)):
                        value = float(cell.value)
                    else:
                        new_warnings.append(
                            Exception(
                                f"{cell.column_letter}{cell.row} : La valeur '{cell.value}' "
                                + f"de la colonne {column_from_index[cell.column]} "
                                + "doit être une valeur numérique"
                            )
                        )

            # Integer case
            elif model_field.get_internal_type() == "IntegerField":
                if cell.value is not None:
                    if isinstance(cell.value, (float, int)):
                        value = int(cell.value)
                    else:
                        new_warnings.append(
                            Exception(
                                f"{cell.column_letter}{cell.row} : La valeur '{cell.value}' "
                                + f"de la colonne {column_from_index[cell.column]} "
                                + "doit être une valeur numérique"
                            )
                        )

            # String case
            elif model_field.get_internal_type() == "CharField":
                if cell.value is not None:
                    if isinstance(cell.value, (float, int, str)):
                        value = cell.value
                    else:
                        new_warnings.append(
                            Exception(
                                f"{cell.column_letter}{cell.row} : La valeur '{cell.value}' "
                                + f"de la colonne {column_from_index[cell.column]} "
                                + "doit être une valeur alphanumeric"
                            )
                        )
        my_row[key] = value

    return my_row, empty_line, new_warnings